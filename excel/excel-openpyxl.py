# _*_ coding: utf-8 _*_
# @Time     : 2020/10/29 13:09
# @Author   : Peter
# @File     : excel-openpyxl.py
# pip install openpyxl

# open a excel file
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, Alignment

border_a = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)

# 如果字体变化，适当调整这两个参数即可
# 一个汉字在Excel中的大致列宽
h_w = 2.1
# 一个英文字符在Excel中的大致列宽
n_w = 1.1
# 姓名列在你的Excel文件中的列标签
label_name = '姓名'

wb_read = openpyxl.load_workbook(filename='./data/openpyxl-data.xlsx')

wb_write = openpyxl.Workbook()
sheet_list = wb_read.sheetnames
wsr = wb_read[sheet_list[0]]


# get where is the label about '姓名'
def get_position_name(label_str=label_name, ws_obj=wsr):
    for i in range(1, ws_obj.max_row + 1):
        for j in range(1, ws_obj.max_column + 1):
            if ws_obj.cell(i, j).value == label_str:
                return i, j
    return 0, 0


# move a row from one sheet object to another
def move_row(s_sheet_obj, s_row_no, d_sheet_obj, d_row_no):
    if s_sheet_obj.max_column >= 1:
        d_sheet_obj.row_dimensions[d_row_no].height = 25
        for i0 in range(1, s_sheet_obj.max_column + 1):
            d_sheet_obj.cell(d_row_no, i0).value = s_sheet_obj.cell(s_row_no, i0).value
        print("successful!!!")
    else:
        print("There is not any data in the source obj!!!")


# set the width of column of one sheet
def set_width(s_s_obj):
    max_col = s_s_obj.max_column
    for i in range(1, max_col + 1):
        width_col = get_max_col_width(s_s_obj[get_column_letter(i)])
        s_s_obj.column_dimensions[get_column_letter(i)].width = width_col


def set_height(s_s_obj, start_r, end_r):
    for i in range(start_r, end_r + 1):
        s_s_obj.row_dimensions[i].height = 25
        for j in range(1, s_s_obj.max_column + 1):
            s_s_obj.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')


# 得到一列中的最大列宽
def get_max_col_width(col_obj):
    length_max = 0
    for ce in col_obj:
        ce = str(ce.value)
        ce_char = count_char(ce)
        lenth_t = int(ce_char[0] * h_w + ce_char[1] * n_w + 0.9)
        if lenth_t > length_max:
            length_max = lenth_t
    return length_max


# 给特定区域内的单元格加上框线
def draw_lines(ss_obj, start_r, start_c, end_r, end_c):
    for i in range(start_r, end_r + 1):
        for j in range(start_c, end_c + 1):
            ss_obj.cell(i, j).border = border_a


# 为了设置列宽的精确，需要知道单元格中有几个汉字几个英文字符
def count_char(s):
    ch_h = 0
    ch_n = 0
    for c in s:
        if ord(c) > 255:
            ch_h = ch_h + 1
        else:
            ch_n = ch_n + 1
    return ch_h, ch_n


name_pos = get_position_name(label_str='姓名')
for n in range(name_pos[0] + 1, wsr.max_row + 1):
    # for n in range(name_pos[0] + 1, 5):
    t_name = wsr.cell(n, name_pos[1]).value
    if n == name_pos[0] + 1:
        wsw = wb_write.active
        wsw.title = t_name
    else:
        wsw = wb_write.create_sheet(title=t_name)
    for i in range(0, len(sheet_list)):
        wsr_temp = wb_read[sheet_list[i]]
        pos_temp = get_position_name(t_name, wsr_temp)
        print(t_name)
        print(pos_temp)
        # write the data of object line into object sheet
        move_row(wsr_temp, pos_temp[0], wsw, 4 * i + 4)
        # write the title data
        move_row(wsr_temp, name_pos[0], wsw, 4 * i + 3)
        # setup the column width
        set_width(wsw)
        draw_lines(wsw, 4 * i + 2, 1, 4 * i + 4, wsw.max_column)
        wsw.merge_cells(start_row=4 * i + 2, start_column=1, end_row=4 * i + 2, end_column=wsw.max_column)
        wsw.cell(4 * i + 2, 1).value = wsr_temp.title
        set_height(wsw, 1, 100)

wb_write.save('./result/openpyxl-openpyxl-ends.xlsx')

