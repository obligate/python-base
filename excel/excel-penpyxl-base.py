# _*_ coding: utf-8 _*_
# @Time     : 2020/10/29 17:12
# @Author   : Peter
# @File     : excel-penpyxl-base.py
import openpyxl
import datetime
from openpyxl.drawing.image import Image

def base_operate():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    print(type(workbook))
    # 获得所有sheet页
    sheet_names = workbook.sheetnames
    print(sheet_names)
    # 获取某个sheet
    stu_sheet = workbook['stu']
    print(stu_sheet)
    print(type(stu_sheet))
    # 复制sheet
    cp_sheet = workbook.copy_worksheet(stu_sheet)
    print(cp_sheet)
    cp_sheet.title = "学籍表"
    print(cp_sheet)

    # cell
    print('============================== cell ===================')
    # 获取当个cell
    cell = stu_sheet['A1']
    print(cell)
    print(type(cell))
    # 获取多个cell
    cells_range = stu_sheet['A1':'C2']
    for cells in cells_range:
        for cell in cells:
            print(cell)
    # 获取指定范围的行
    row = stu_sheet.iter_rows(min_row=1, max_col=3, max_row=1)
    for cell in row:
        print(cell)
    # 获取全部行
    for row in stu_sheet.rows:
        print(row)

    # 获取cell值
    print('============================== cell value ===================')
    for cells in cells_range:
        for cell in cells:
            print(cell.value)


def write_operate():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet('write_operate')
    sheet['A1'] = 'write_666'
    print(sheet['A1'].value)
    # 设置行高
    sheet.row_dimensions[1].height = 50
    # 设置列高
    sheet.column_dimensions['A'].width = 30
    sheet['A2'] = datetime.datetime(2010, 7, 21)
    print(sheet['A2'].value)
    sheet['A3'] = '=SUM(1, 1)'
    print(sheet['A3'].value)    # =SUM(1, 1)   实际中A3单元格值为2
    # 设置行高
    sheet.row_dimensions[1].height = 50
    # 设置列高
    sheet.column_dimensions['A'].width = 30
    # 合并单元格
    sheet.merge_cells('A2:D2')
    # 取消合并单元格
    # sheet.unmerge_cells('A2:D2')
    # 保存
    workbook.save(path)

def insert_image():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet('insert_image')
    # 设置图像
    img = Image(r'./data/01.jpg')
    # img.width, img.height这两个属性分别是对应添加图片的宽高
    newsize = (90, 90)
    img.width, img.height = newsize
    # 设置图像单元格说明
    sheet['A1'] = 'you are my angel'
    # 插入图片
    sheet.add_image(img, 'A1')
    # # 移除sheet
    # workbook.remove(sheet)
    # 保存
    workbook.save(path)

def remove_image():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['insert_image']
    # 移除sheet
    workbook.remove(sheet)
    # 保存
    workbook.save(path)

def update_sheet():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['write_operate']
    val = sheet['A1'].value
    print(val)
    new_val = sheet['A1'].value = 'update A1'
    print(new_val)
    workbook.save(path)

def hidden_sheet_col_row():
    path = r'./data/pyxl-base.xlsx'
    # 加载workbook
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['write_operate']
    # 隐藏行 A-B
    sheet.column_dimensions.group('A', 'B', hidden=True)
    # 隐藏 列 1 -5
    sheet.row_dimensions.group(1, 5, hidden=True)
    workbook.save(path)

# base_operate()
insert_image()
# remove_image()
# update_sheet()
# hidden_sheet_col_row()